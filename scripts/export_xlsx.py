#!/usr/bin/env python3
"""Export the dock to the classic 6-tab styled workbook (GovTech_Company_Tracker.xlsx).

Reads data/companies.json + data/schema.json; writes exports/GovTech_Company_Tracker.xlsx.
Layout matches the original Cowork spreadsheet: one filterable table per sector tab,
Arial, banded rows, category dropdown, bold-green hiring column.
"""
from __future__ import annotations

import datetime as dt
import json
import pathlib

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = ROOT / "exports"

FONT = "Arial"
HEADERS = ["Company", "Location", "Year Founded", "Category", "Description", None]  # [5] set at runtime
STATUS_COLOR = {"Yes": "1E8449", "Sales (non-AE)": "B9770E",
                "None found": "7F8C8D", "Unknown": "7F8C8D"}
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def row_values(comp: dict) -> list:
    """One spreadsheet row. Optional fields render blank rather than crashing
    the run: this exporter once killed a 41-minute refresh over a missing
    year_founded, taking the day's uncommitted data with it. Only the fields
    selftest's validate() actually guarantees may be indexed directly."""
    h = comp["hiring"]
    hire_text = h["status"] + (f" - {h['note']}" if h.get("note") else "")
    return [comp["name"], comp.get("location", ""), comp.get("year_founded"),
            comp["category"], comp.get("description", ""), hire_text]


def main() -> None:
    companies = json.load(open(DATA / "companies.json"))
    schema = json.load(open(DATA / "schema.json"))
    meta = json.load(open(DATA / "meta.json"))
    stamp = dt.date.fromisoformat(meta["last_run"]).strftime("%-m/%-d/%y") \
        if meta.get("last_run") else "never"
    headers = HEADERS[:5] + [f"Hiring AEs? ({stamp})"]

    OUT.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sec in schema["sectors"]:
        head, band = sec["xlsx"]["head"], sec["xlsx"]["band"]
        ws = wb.create_sheet(sec["name"])
        ws.sheet_properties.tabColor = head

        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=head)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BORDER
        ws.row_dimensions[1].height = 20

        rows = [c for c in companies if c["sector"] == sec["name"]]
        rows.sort(key=lambda c: sec["categories"].index(c["category"]))  # stable: keeps file order within category
        r = 2
        for comp in rows:
            values = row_values(comp)
            h = comp["hiring"]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = BORDER
                if c == 3:
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="center")
                if c == 6:
                    cell.font = Font(name=FONT, size=11,
                                     color=STATUS_COLOR.get(h["status"], "7F8C8D"),
                                     bold=(h["status"] == "Yes"))
                else:
                    cell.font = Font(name=FONT, size=11)
                if r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=band)
            r += 1

        for i, w in enumerate([34, 26, 14, 24, 78, 34], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.auto_filter.ref = f"A1:F{r - 1}"
        ws.freeze_panes = "A2"

        dv = DataValidation(type="list",
                            formula1='"' + ",".join(sec["categories"]) + '"',
                            allow_blank=True)
        dv.error = "Pick one of: " + ", ".join(sec["categories"])
        dv.errorTitle = "Invalid category"
        ws.add_data_validation(dv)
        dv.add("D2:D500")
        ws.cell(row=1, column=1).comment = Comment(
            "Exported from the govtech-dock repo - edit data/companies.json and re-run "
            "scripts/export_xlsx.py rather than editing this file.", "govtech-dock")

    path = OUT / "GovTech_Company_Tracker.xlsx"
    wb.save(path)
    print(f"wrote {path.relative_to(ROOT)} ({len(companies)} companies, "
          f"{len(schema['sectors'])} tabs)")


if __name__ == "__main__":
    main()
